import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.schema import ExtractionResult

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FONT = Font(name="游ゴシック", size=9, bold=True)
CELL_FONT = Font(name="游ゴシック", size=9)
TITLE_FONT = Font(name="游ゴシック", size=14, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _format_date(d: date | None) -> str:
    if d is None:
        return ""
    return f"令和{d.year - 2018}年{d.month}月{d.day}日"


def _format_amount(amount: int | None) -> str:
    if amount is None:
        return ""
    return f"{amount:,}"


def _set_cell(ws, row, col, value, font=None, alignment=None, border=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if fill:
        cell.fill = fill
    return cell


def _build_honpyo(wb: Workbook, data: ExtractionResult):
    """本表シートを構築"""
    ws = wb.active
    ws.title = "本表"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    _set_cell(ws, row, 1, "事前確定届出給与に関する届出書（本表）", TITLE_FONT, CENTER)
    row += 2

    corp = data.corporation
    res = data.resolution

    corp_fields = [
        ("法人名", corp.name),
        ("法人番号", corp.corporation_number or ""),
        ("納税地", f"〒{corp.postal_code or ''} {corp.address}"),
        ("電話番号", corp.phone or ""),
        ("代表者氏名", corp.representative or ""),
        ("事業年度", f"{_format_date(corp.fiscal_year_start)} ～ {_format_date(corp.fiscal_year_end)}"),
        ("資本金", f"{_format_amount(corp.capital)}円" if corp.capital else ""),
        ("所轄税務署", f"{corp.tax_office}長 殿"),
    ]

    for label, value in corp_fields:
        _set_cell(ws, row, 1, label, HEADER_FONT, LEFT, THIN_BORDER, HEADER_FILL)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _set_cell(ws, row, 3, value, CELL_FONT, LEFT, THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    _set_cell(ws, row, 1, "届出の事由等", HEADER_FONT, CENTER, THIN_BORDER, HEADER_FILL)
    row += 1

    res_fields = [
        ("① 決議をした日", _format_date(res.decision_date)),
        ("① 決議をした機関等", res.decision_body),
        ("② 職務執行開始日", _format_date(res.execution_start_date)),
        ("③ 当該事業年度開始の日", _format_date(res.fiscal_year_basis)),
        ("届出期限の根拠区分", res.filing_deadline_basis or "イ"),
        ("⑤ 定期同額給与としない理由", res.reason_for_bonus_timing or ""),
        ("⑥ その他参考事項", res.remarks or ""),
    ]

    for label, value in res_fields:
        _set_cell(ws, row, 1, label, HEADER_FONT, LEFT, THIN_BORDER, HEADER_FILL)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        _set_cell(ws, row, 4, value, CELL_FONT, LEFT, THIN_BORDER)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=10)
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    _set_cell(ws, row, 1, "④ 事前確定届出給与等の状況（概要）", HEADER_FONT, CENTER, THIN_BORDER, HEADER_FILL)
    row += 1

    headers = ["氏名", "役職名", "支給時期", "支給額（円）", "定期同額給与月額"]
    for i, h in enumerate(headers):
        col = i * 2 + 1
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        _set_cell(ws, row, col, h, HEADER_FONT, CENTER, THIN_BORDER, HEADER_FILL)
        ws.cell(row=row, column=col + 1).border = THIN_BORDER
    row += 1

    for officer in data.officers:
        for j, payment in enumerate(officer.payments):
            name_val = officer.name if j == 0 else ""
            pos_val = officer.position if j == 0 else ""
            reg_val = _format_amount(officer.regular_payment) if j == 0 and officer.regular_payment else ""

            vals = [name_val, pos_val, _format_date(payment.payment_date), _format_amount(payment.amount), reg_val]
            for i, v in enumerate(vals):
                col = i * 2 + 1
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
                _set_cell(ws, row, col, v, CELL_FONT, CENTER, THIN_BORDER)
                ws.cell(row=row, column=col + 1).border = THIN_BORDER
            row += 1


def _build_futahyo(wb: Workbook, data: ExtractionResult):
    """付表1シートを構築（対象役員ごとの明細）"""
    ws = wb.create_sheet("付表1")
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    _set_cell(ws, row, 1, "事前確定届出給与等の状況（付表1）", TITLE_FONT, CENTER)
    row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set_cell(ws, row, 1, "法人名", HEADER_FONT, LEFT, THIN_BORDER, HEADER_FILL)
    _set_cell(ws, row, 4, data.corporation.name, CELL_FONT, LEFT, THIN_BORDER)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=10)
    for col in range(1, 11):
        ws.cell(row=row, column=col).border = THIN_BORDER
    row += 2

    for idx, officer in enumerate(data.officers):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        _set_cell(ws, row, 1, f"【{idx + 1}】 {officer.name}（{officer.position}）",
                  Font(name="游ゴシック", size=11, bold=True), LEFT, THIN_BORDER, HEADER_FILL)
        row += 1

        info_fields = [
            ("氏名フリガナ", officer.name_kana or ""),
            ("就任年月日", _format_date(officer.appointment_date)),
            ("区分", officer.category or "事前確定届出給与"),
        ]
        for label, value in info_fields:
            _set_cell(ws, row, 1, label, HEADER_FONT, LEFT, THIN_BORDER, HEADER_FILL)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            _set_cell(ws, row, 3, value, CELL_FONT, LEFT, THIN_BORDER)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        pay_headers = ["回", "支給時期", "届出額（円）", "届出額（税込）", "備考"]
        for i, h in enumerate(pay_headers):
            col = i * 2 + 1
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            _set_cell(ws, row, col, h, HEADER_FONT, CENTER, THIN_BORDER, HEADER_FILL)
            ws.cell(row=row, column=col + 1).border = THIN_BORDER
        row += 1

        for j, payment in enumerate(officer.payments):
            vals = [
                str(j + 1),
                _format_date(payment.payment_date),
                _format_amount(payment.amount),
                _format_amount(payment.amount),
                payment.memo or "",
            ]
            for i, v in enumerate(vals):
                col = i * 2 + 1
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
                _set_cell(ws, row, col, v, CELL_FONT, CENTER, THIN_BORDER)
                ws.cell(row=row, column=col + 1).border = THIN_BORDER
            row += 1

        if officer.regular_payment:
            _set_cell(ws, row, 1, "定期同額給与（月額）", HEADER_FONT, LEFT, THIN_BORDER, HEADER_FILL)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            _set_cell(ws, row, 5, f"{_format_amount(officer.regular_payment)}円", CELL_FONT, LEFT, THIN_BORDER)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=10)
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        row += 1


def build_xlsx(data: ExtractionResult) -> bytes:
    """ExtractionResult から Excel バイナリを生成"""
    wb = Workbook()
    _build_honpyo(wb, data)
    _build_futahyo(wb, data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
